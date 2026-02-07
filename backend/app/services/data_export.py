"""
Data Export Service
Export data to CSV, Excel, JSON formats
"""
import io
import csv
import json
from datetime import datetime
from typing import Optional, Dict, List, Any
from dataclasses import dataclass


@dataclass
class ExportResult:
    """Export result."""
    filename: str
    content_type: str
    data: bytes
    record_count: int


class DataExportService:
    """
    Export data in various formats.
    
    Formats supported:
    - CSV
    - JSON
    - Excel (XLSX) if openpyxl available
    """
    
    def __init__(self):
        self._check_excel_support()
    
    def _check_excel_support(self):
        try:
            import openpyxl
            self.excel_available = True
        except ImportError:
            self.excel_available = False
    
    def export_projects(
        self,
        projects: List[Dict],
        format: str = "csv"
    ) -> ExportResult:
        """Export projects data."""
        columns = [
            "id", "name", "owner", "address", "city", "state", "pin_code",
            "roof_area_sqm", "building_type", "status", "tank_capacity",
            "yearly_yield", "estimated_cost", "created_at"
        ]
        
        return self._export(projects, columns, "projects", format)
    
    def export_assessments(
        self,
        assessments: List[Dict],
        format: str = "csv"
    ) -> ExportResult:
        """Export assessments data."""
        columns = [
            "id", "project_id", "project_name", "roof_area_sqm",
            "annual_rainfall_mm", "runoff_coefficient", "yearly_yield_liters",
            "recommended_tank_liters", "estimated_cost_min", "estimated_cost_max",
            "roi_years", "carbon_offset_kg", "created_at"
        ]
        
        return self._export(assessments, columns, "assessments", format)
    
    def export_installers(
        self,
        installers: List[Dict],
        format: str = "csv"
    ) -> ExportResult:
        """Export installers data."""
        columns = [
            "id", "company_name", "contact_person", "email", "phone",
            "license_number", "gst_number", "is_verified", "rating",
            "projects_completed", "service_areas", "created_at"
        ]
        
        return self._export(installers, columns, "installers", format)
    
    def export_payments(
        self,
        payments: List[Dict],
        format: str = "csv"
    ) -> ExportResult:
        """Export payments data."""
        columns = [
            "id", "project_id", "project_name", "user_name", "amount",
            "payment_type", "status", "gateway", "gateway_payment_id",
            "utr_number", "created_at", "completed_at"
        ]
        
        return self._export(payments, columns, "payments", format)
    
    def export_verifications(
        self,
        verifications: List[Dict],
        format: str = "csv"
    ) -> ExportResult:
        """Export verifications data."""
        columns = [
            "id", "project_id", "project_name", "submission_type",
            "status", "submitted_by", "reviewed_by", "fraud_score",
            "submitted_at", "reviewed_at", "rejection_reason"
        ]
        
        return self._export(verifications, columns, "verifications", format)
    
    def export_sensor_readings(
        self,
        readings: List[Dict],
        format: str = "csv"
    ) -> ExportResult:
        """Export sensor readings data."""
        columns = [
            "timestamp", "sensor_id", "sensor_type", "project_id",
            "value", "unit", "battery_percent"
        ]
        
        return self._export(readings, columns, "sensor_readings", format)
    
    def export_analytics(
        self,
        analytics: Dict,
        format: str = "json"
    ) -> ExportResult:
        """Export analytics data (JSON only)."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"analytics_{timestamp}.json"
        
        data = json.dumps(analytics, indent=2, default=str).encode()
        
        return ExportResult(
            filename=filename,
            content_type="application/json",
            data=data,
            record_count=len(analytics)
        )
    
    def export_custom(
        self,
        data: List[Dict],
        columns: List[str],
        name: str,
        format: str = "csv"
    ) -> ExportResult:
        """Export custom data with specified columns."""
        return self._export(data, columns, name, format)
    
    def _export(
        self,
        data: List[Dict],
        columns: List[str],
        name: str,
        format: str
    ) -> ExportResult:
        """Internal export method."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format == "csv":
            return self._to_csv(data, columns, f"{name}_{timestamp}.csv")
        elif format == "json":
            return self._to_json(data, f"{name}_{timestamp}.json")
        elif format == "xlsx" and self.excel_available:
            return self._to_excel(data, columns, f"{name}_{timestamp}.xlsx")
        else:
            return self._to_csv(data, columns, f"{name}_{timestamp}.csv")
    
    def _to_csv(
        self,
        data: List[Dict],
        columns: List[str],
        filename: str
    ) -> ExportResult:
        """Export to CSV format."""
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction='ignore')
        
        writer.writeheader()
        for row in data:
            # Flatten nested dicts and convert dates
            flat_row = {}
            for col in columns:
                value = row.get(col, "")
                if isinstance(value, datetime):
                    value = value.isoformat()
                elif isinstance(value, (list, dict)):
                    value = json.dumps(value)
                flat_row[col] = value
            writer.writerow(flat_row)
        
        content = buffer.getvalue().encode('utf-8-sig')  # BOM for Excel compatibility
        
        return ExportResult(
            filename=filename,
            content_type="text/csv",
            data=content,
            record_count=len(data)
        )
    
    def _to_json(
        self,
        data: List[Dict],
        filename: str
    ) -> ExportResult:
        """Export to JSON format."""
        content = json.dumps({
            "exported_at": datetime.now().isoformat(),
            "count": len(data),
            "data": data
        }, indent=2, default=str).encode()
        
        return ExportResult(
            filename=filename,
            content_type="application/json",
            data=content,
            record_count=len(data)
        )
    
    def _to_excel(
        self,
        data: List[Dict],
        columns: List[str],
        filename: str
    ) -> ExportResult:
        """Export to Excel format."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Header styling
        header_fill = PatternFill(start_color="0ea5e9", end_color="0ea5e9", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_idx, row in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row.get(col_name, "")
                if isinstance(value, datetime):
                    value = value.isoformat()
                elif isinstance(value, (list, dict)):
                    value = json.dumps(value)
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(col_name)
            for row in data[:100]:  # Check first 100 rows
                value = str(row.get(col_name, ""))
                max_length = max(max_length, len(value))
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 50)
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        
        return ExportResult(
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            data=content,
            record_count=len(data)
        )


# Singleton
_export_service: Optional[DataExportService] = None

def get_export_service() -> DataExportService:
    global _export_service
    if _export_service is None:
        _export_service = DataExportService()
    return _export_service

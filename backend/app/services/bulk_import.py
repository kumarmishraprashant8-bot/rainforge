"""
Bulk Import Service
Import projects and data from CSV/Excel files
"""
import io
import csv
from datetime import datetime
from typing import Optional, Dict, List, Any, Tuple
from dataclasses import dataclass
import re


@dataclass
class ImportResult:
    """Import result."""
    success_count: int
    error_count: int
    errors: List[Dict[str, Any]]
    imported_ids: List[Any]


@dataclass
class ValidationError:
    """Validation error."""
    row: int
    column: str
    value: Any
    message: str


class BulkImportService:
    """
    Bulk import data from CSV/Excel files.
    
    Supports:
    - Project bulk import
    - Sensor registration
    - User import
    - Historical data import
    """
    
    # Column mappings (CSV header -> internal field)
    PROJECT_COLUMNS = {
        "project_name": "name",
        "name": "name",
        "owner_name": "owner_name",
        "owner": "owner_name",
        "address": "address",
        "city": "city",
        "state": "state",
        "pin_code": "pin_code",
        "pincode": "pin_code",
        "roof_area": "roof_area_sqm",
        "roof_area_sqm": "roof_area_sqm",
        "roof_area_sqft": "roof_area_sqft",
        "building_type": "building_type",
        "type": "building_type",
        "phone": "phone",
        "email": "email",
        "latitude": "latitude",
        "lat": "latitude",
        "longitude": "longitude",
        "lng": "longitude",
        "lon": "longitude"
    }
    
    SENSOR_COLUMNS = {
        "sensor_id": "sensor_id",
        "project_id": "project_id",
        "sensor_type": "sensor_type",
        "type": "sensor_type",
        "dev_eui": "dev_eui",
        "protocol": "protocol"
    }
    
    def __init__(self):
        pass
    
    async def import_projects(
        self,
        file_content: bytes,
        file_type: str = "csv",
        user_id: str = None,
        tenant_id: str = "default"
    ) -> ImportResult:
        """Import projects from CSV/Excel file."""
        rows = self._parse_file(file_content, file_type)
        
        if not rows:
            return ImportResult(
                success_count=0,
                error_count=0,
                errors=[{"message": "No data found in file"}],
                imported_ids=[]
            )
        
        # Map headers
        headers = [h.lower().strip() for h in rows[0]]
        column_map = {}
        for idx, header in enumerate(headers):
            if header in self.PROJECT_COLUMNS:
                column_map[idx] = self.PROJECT_COLUMNS[header]
        
        imported_ids = []
        errors = []
        
        for row_num, row in enumerate(rows[1:], 2):
            try:
                # Parse row
                project_data = {}
                for idx, value in enumerate(row):
                    if idx in column_map:
                        field = column_map[idx]
                        project_data[field] = self._clean_value(value, field)
                
                # Validate
                validation_errors = self._validate_project(project_data, row_num)
                if validation_errors:
                    errors.extend(validation_errors)
                    continue
                
                # Convert sqft to sqm if needed
                if "roof_area_sqft" in project_data:
                    project_data["roof_area_sqm"] = project_data["roof_area_sqft"] * 0.0929
                    del project_data["roof_area_sqft"]
                
                # Add metadata
                project_data["user_id"] = user_id
                project_data["tenant_id"] = tenant_id
                project_data["status"] = "imported"
                project_data["created_at"] = datetime.utcnow()
                
                # In production, would save to database
                # For now, generate mock ID
                project_id = f"IMP-{row_num}-{datetime.now().timestamp():.0f}"
                imported_ids.append(project_id)
                
            except Exception as e:
                errors.append({
                    "row": row_num,
                    "message": str(e)
                })
        
        return ImportResult(
            success_count=len(imported_ids),
            error_count=len(errors),
            errors=errors[:50],  # Limit error list
            imported_ids=imported_ids
        )
    
    async def import_sensors(
        self,
        file_content: bytes,
        file_type: str = "csv"
    ) -> ImportResult:
        """Import sensors from CSV file."""
        rows = self._parse_file(file_content, file_type)
        
        if not rows:
            return ImportResult(0, 0, [{"message": "No data"}], [])
        
        headers = [h.lower().strip() for h in rows[0]]
        column_map = {}
        for idx, header in enumerate(headers):
            if header in self.SENSOR_COLUMNS:
                column_map[idx] = self.SENSOR_COLUMNS[header]
        
        imported_ids = []
        errors = []
        
        for row_num, row in enumerate(rows[1:], 2):
            try:
                sensor_data = {}
                for idx, value in enumerate(row):
                    if idx in column_map:
                        sensor_data[column_map[idx]] = value.strip()
                
                # Validate required fields
                if not sensor_data.get("sensor_id"):
                    errors.append({"row": row_num, "message": "Missing sensor_id"})
                    continue
                
                if not sensor_data.get("project_id"):
                    errors.append({"row": row_num, "message": "Missing project_id"})
                    continue
                
                imported_ids.append(sensor_data["sensor_id"])
                
            except Exception as e:
                errors.append({"row": row_num, "message": str(e)})
        
        return ImportResult(
            success_count=len(imported_ids),
            error_count=len(errors),
            errors=errors[:50],
            imported_ids=imported_ids
        )
    
    async def import_historical_readings(
        self,
        file_content: bytes,
        sensor_id: str,
        file_type: str = "csv"
    ) -> ImportResult:
        """Import historical sensor readings."""
        rows = self._parse_file(file_content, file_type)
        
        if not rows:
            return ImportResult(0, 0, [{"message": "No data"}], [])
        
        imported_count = 0
        errors = []
        
        for row_num, row in enumerate(rows[1:], 2):
            try:
                if len(row) < 2:
                    continue
                
                timestamp = self._parse_date(row[0])
                value = float(row[1])
                
                if timestamp and value is not None:
                    imported_count += 1
                else:
                    errors.append({"row": row_num, "message": "Invalid data"})
                    
            except Exception as e:
                errors.append({"row": row_num, "message": str(e)})
        
        return ImportResult(
            success_count=imported_count,
            error_count=len(errors),
            errors=errors[:50],
            imported_ids=[]
        )
    
    def get_template(self, import_type: str) -> Tuple[str, bytes]:
        """Get CSV template for import type."""
        templates = {
            "projects": [
                "project_name,owner_name,address,city,state,pin_code,roof_area_sqm,building_type,phone,email",
                "Sample Project,John Doe,123 Main St,Mumbai,Maharashtra,400001,100,residential,9876543210,john@example.com"
            ],
            "sensors": [
                "sensor_id,project_id,sensor_type,protocol,dev_eui",
                "SENSOR001,1,tank_level,mqtt,",
                "LORA001,2,tank_level,lorawan,00:11:22:33:44:55:66:77"
            ],
            "readings": [
                "timestamp,value,unit",
                "2024-01-15 10:00:00,75.5,percent",
                "2024-01-15 11:00:00,74.2,percent"
            ]
        }
        
        content = templates.get(import_type, templates["projects"])
        csv_content = "\n".join(content).encode('utf-8-sig')
        
        return f"{import_type}_template.csv", csv_content
    
    def _parse_file(
        self,
        content: bytes,
        file_type: str
    ) -> List[List[str]]:
        """Parse file content to rows."""
        if file_type in ["csv", "text/csv"]:
            return self._parse_csv(content)
        elif file_type in ["xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]:
            return self._parse_excel(content)
        else:
            return self._parse_csv(content)
    
    def _parse_csv(self, content: bytes) -> List[List[str]]:
        """Parse CSV content."""
        # Try different encodings
        for encoding in ['utf-8-sig', 'utf-8', 'latin-1']:
            try:
                text = content.decode(encoding)
                reader = csv.reader(io.StringIO(text))
                return list(reader)
            except:
                continue
        return []
    
    def _parse_excel(self, content: bytes) -> List[List[str]]:
        """Parse Excel content."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            return rows
        except ImportError:
            return []
        except Exception:
            return []
    
    def _clean_value(self, value: Any, field: str) -> Any:
        """Clean and convert value based on field type."""
        if value is None:
            return None
        
        value = str(value).strip()
        
        if not value:
            return None
        
        # Numeric fields
        if field in ["roof_area_sqm", "roof_area_sqft", "latitude", "longitude"]:
            try:
                return float(value.replace(",", ""))
            except:
                return None
        
        # Phone number
        if field == "phone":
            return re.sub(r'[^\d+]', '', value)
        
        # Pin code
        if field == "pin_code":
            return re.sub(r'[^\d]', '', value)[:6]
        
        return value
    
    def _validate_project(self, data: Dict, row_num: int) -> List[Dict]:
        """Validate project data."""
        errors = []
        
        # Required fields
        if not data.get("name"):
            errors.append({"row": row_num, "column": "name", "message": "Project name is required"})
        
        if not data.get("city"):
            errors.append({"row": row_num, "column": "city", "message": "City is required"})
        
        # Validate roof area
        roof_area = data.get("roof_area_sqm") or data.get("roof_area_sqft")
        if roof_area is not None:
            if roof_area <= 0:
                errors.append({"row": row_num, "column": "roof_area", "message": "Roof area must be positive"})
            if roof_area > 100000:
                errors.append({"row": row_num, "column": "roof_area", "message": "Roof area seems too large"})
        
        # Validate pin code
        pin = data.get("pin_code")
        if pin and len(pin) != 6:
            errors.append({"row": row_num, "column": "pin_code", "message": "Pin code must be 6 digits"})
        
        # Validate coordinates
        lat = data.get("latitude")
        if lat is not None and (lat < -90 or lat > 90):
            errors.append({"row": row_num, "column": "latitude", "message": "Invalid latitude"})
        
        lng = data.get("longitude")
        if lng is not None and (lng < -180 or lng > 180):
            errors.append({"row": row_num, "column": "longitude", "message": "Invalid longitude"})
        
        return errors
    
    def _parse_date(self, value: str) -> Optional[datetime]:
        """Parse date string to datetime."""
        formats = [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y",
            "%d-%m-%Y"
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(value.strip(), fmt)
            except:
                continue
        return None


# Singleton
_import_service: Optional[BulkImportService] = None

def get_import_service() -> BulkImportService:
    global _import_service
    if _import_service is None:
        _import_service = BulkImportService()
    return _import_service
